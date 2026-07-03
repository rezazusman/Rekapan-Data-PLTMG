import os
import json
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import pandas as pd
import shutil
import tempfile

JSON_FILE = "latest_data.json"

HISTORY_ROOT = "History"

HEADER = [
    "Timestamp",
    "UNIT1_MW","UNIT1_MVAR",
    "UNIT2_MW","UNIT2_MVAR",
    "UNIT3_MW","UNIT3_MVAR",
    "UNIT4_MW","UNIT4_MVAR",
    "UNIT5_MW","UNIT5_MVAR"
]

last_timestamp = ""

current_date = ""
current_file = ""

wb = None
ws = None

# workbook hari sebelumnya
wb_old = None
ws_old = None

pending_close = False
pending_file = ""
pending_time = 0

# save workbook aktif setiap 30 detik
SAVE_INTERVAL = 30
last_save = time.time()


# ===========================================================
# Membuka / membuat file history
# ===========================================================

def buka_file_history(tanggal):

    global wb
    global ws
    global current_file

    global wb_old
    global ws_old
    global pending_close
    global pending_file
    global pending_time

    # ======================================================
    # Jangan langsung save workbook lama.
    # Beri waktu 10 detik.
    # ======================================================

    if wb is not None and not pending_close:

        wb_old = wb
        ws_old = ws

        pending_close = True
        pending_file = current_file
        pending_time = time.time()

        print(f"Workbook lama menunggu disimpan : {pending_file}")

    dt = datetime.strptime(tanggal,"%Y-%m-%d")

    tahun = str(dt.year)
    bulan = dt.strftime("%B")

    folder = os.path.join(
        HISTORY_ROOT,
        tahun,
        bulan
    )

    os.makedirs(folder, exist_ok=True)

    current_file = os.path.join(
        folder,
        f"LOG_{tanggal}.xlsx"
    )

    if os.path.exists(current_file):

        wb = load_workbook(current_file)
        ws = wb.active

    else:

        wb = Workbook()
        ws = wb.active
        ws.title = "LOG"

        ws.append(HEADER)

        save_workbook_atomic(wb, current_file)

    print(f"History aktif : {current_file}")

def backup_to_csv(xlsx_file):

    try:

        csv_file = xlsx_file.replace(".xlsx", ".csv")

        df = pd.read_excel(xlsx_file)

        df.to_csv(
            csv_file,
            index=False,
            encoding="utf-8-sig"
        )

        print(f"Backup CSV berhasil : {csv_file}")

    except Exception as e:

        print(f"Gagal membuat CSV : {e}")


def backup_to_bak(xlsx_file):

    try:

        bak_file = xlsx_file.replace(".xlsx", ".bak.xlsx")

        shutil.copy2(xlsx_file, bak_file)

        print(f"Backup XLSX berhasil : {bak_file}")

    except Exception as e:

        print(f"Gagal membuat backup XLSX : {e}")

def save_workbook_atomic(workbook, filename):
    """
    Menyimpan workbook dengan metode atomic.
    Menghindari file xlsx korup jika proses terhenti saat save.
    """

    try:

        folder = os.path.dirname(filename)

        fd, temp_file = tempfile.mkstemp(
            suffix=".tmp.xlsx",
            dir=folder
        )

        os.close(fd)

        workbook.save(temp_file)

        os.replace(temp_file, filename)

        return True

    except Exception as e:

        print("Atomic Save Error :", e)

        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except:
            pass

        return False
    
# ===========================================================
# MAIN
# ===========================================================

print("="*50)
print("History Writer")
print("="*50)

last_save = time.time()

try:

    while True:

        try:

            if not os.path.exists(JSON_FILE):
                time.sleep(1)
                continue

            with open(JSON_FILE,"r") as f:
                obj = json.load(f)

            timestamp = obj["timestamp"]

            # ===================================================
            # Save workbook lama setelah 10 detik
            # ===================================================

            if pending_close:

                if time.time() - pending_time >= 10:

                    try:

                        print(f"Menyimpan workbook lama : {pending_file}")

                        if save_workbook_atomic(wb_old, pending_file):

                            wb_old.close()

                            backup_to_csv(pending_file)

                            backup_to_bak(pending_file)
                            print(f"Workbook lama selesai disimpan.")

                    except Exception as e:

                        print(e)

                    wb_old = None
                    ws_old = None

                    pending_close = False

            # ===================================================
            # Tidak ada data baru
            # ===================================================

            if timestamp == last_timestamp:
                time.sleep(0.2)
                continue

            last_timestamp = timestamp

            tanggal = timestamp[:10]

            # ===================================================
            # Ganti hari
            # ===================================================

            if tanggal != current_date:

                current_date = tanggal

                buka_file_history(current_date)

                last_save = time.time()

            data = obj["data"]

            ws.append([

                timestamp,

                data.get("UNIT1_MW",""),
                data.get("UNIT1_MVAR",""),

                data.get("UNIT2_MW",""),
                data.get("UNIT2_MVAR",""),

                data.get("UNIT3_MW",""),
                data.get("UNIT3_MVAR",""),

                data.get("UNIT4_MW",""),
                data.get("UNIT4_MVAR",""),

                data.get("UNIT5_MW",""),
                data.get("UNIT5_MVAR",""),

            ])

            # ===================================================
            # Save workbook aktif setiap 30 detik
            # ===================================================

            if time.time() - last_save >= SAVE_INTERVAL:

                try:
                    if save_workbook_atomic(wb, current_file):

                        last_save = time.time()    

                except Exception as e:
                    print("Gagal save workbook aktif :", e)

            print(timestamp)

        except Exception as e:

            print(e)

        time.sleep(0.2)

except KeyboardInterrupt:

    print("\nHistory Writer dihentikan.")

finally:

    print("Menyimpan workbook terakhir...")

    try:

        if wb is not None:

            if save_workbook_atomic(wb, current_file):

                wb.close()

                backup_to_csv(current_file)

                backup_to_bak(current_file)
    except Exception as e:

        print(e)

    try:

        if wb_old is not None:

            if save_workbook_atomic(wb_old, pending_file):

                wb_old.close()

                backup_to_csv(pending_file)

                backup_to_bak(pending_file)

    except Exception as e:

        print(e)

    print("History Writer selesai.")