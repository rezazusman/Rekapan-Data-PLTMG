import cv2
import yt_dlp
import easyocr
import json
import time
import re
import os
from openpyxl import Workbook, load_workbook
from threading import Thread, Lock


# ==========================================================
# KONFIGURASI
# ==========================================================

CHANNEL_URL = "https://www.youtube.com/@LUWUK40MW-l7p/live"

CHECK_LIVE_INTERVAL = 60      # cek live baru setiap 60 detik
RECONNECT_INTERVAL = 1800      # reconnect stream tiap 5 menit

EXCEL_FILE = "MW_MVAR_LOG.xlsx"
JSON_FILE = "latest_data.json"

INTERVAL = 0.2      # detik

# ==========================================================
# VALIDASI NILAI OCR
# ==========================================================

MAX_MW = 9.3
MAX_MVAR = 1.5

# ==========================================================
# LOAD ROI
# ==========================================================

with open("roi.json", "r") as f:
    ROIS = json.load(f)

print("====================================")
print("ROI berhasil dimuat.")
print("Jumlah ROI :", len(ROIS))
print("====================================")

# ==========================================================
# EASYOCR
# ==========================================================

print("Memuat EasyOCR...")

reader = easyocr.Reader(
    ['en'],
    gpu=False
)

print("EasyOCR siap.")

# ==========================================================
# FUNGSI MENCARI LIVE TERBARU
# ==========================================================

def cari_live():

    print("Mencari live terbaru...")

    ydl_opts = {
        "quiet": True,
        "format": "best",
        "extract_flat": False
    }

    try:

        with yt_dlp.YoutubeDL(ydl_opts) as ydl:

            info = ydl.extract_info(
                CHANNEL_URL,
                download=False
            )

        if info.get("is_live"):

            stream = info.get("url")

            if stream:

                print("=================================")
                print("LIVE DITEMUKAN")
                print(info.get("title"))
                print("=================================")

                return stream

        print("Channel belum sedang live.")
        return None

    except Exception as e:

        print(f"Gagal mengambil live: {e}")
        return None

# ==========================================================
# MEMBUKA YOUTUBE LIVE
# ==========================================================

STREAM_URL = None

while STREAM_URL is None:

    STREAM_URL = cari_live()

    if STREAM_URL is None:

        print("Menunggu live...")

        time.sleep(CHECK_LIVE_INTERVAL)

cap = cv2.VideoCapture(STREAM_URL)

cap.set(cv2.CAP_PROP_BUFFERSIZE,1)

# ==========================================
# Background Frame Reader
# ==========================================

latest_frame = None
frame_lock = Lock()

def frame_reader():

    global latest_frame

    while True:

        if cap.isOpened():

            ret, frame = cap.read()

            if ret:

                with frame_lock:
                    latest_frame = frame

            else:
                time.sleep(0.01)

        else:
            time.sleep(0.1)

Thread(
    target=frame_reader,
    daemon=True
).start()

print("Streaming berhasil dibuka.")

if not cap.isOpened():

    print("Tidak dapat membuka video.")
    exit()

print("YouTube Live berhasil dibuka.")

# ==========================================================
# MEMBUAT FILE EXCEL
# ==========================================================

if not os.path.exists(EXCEL_FILE):

    wb = Workbook()

    ws = wb.active

    ws.title = "LOG"

    header = [
        "Timestamp",
        "UNIT1_MW",
        "UNIT1_MVAR",
        "UNIT2_MW",
        "UNIT2_MVAR",
        "UNIT3_MW",
        "UNIT3_MVAR",
        "UNIT4_MW",
        "UNIT4_MVAR",
        "UNIT5_MW",
        "UNIT5_MVAR"
    ]

    ws.append(header)

    wb.save(EXCEL_FILE)

print("Excel siap.")

# ==========================================================
# FUNGSI OCR
# ==========================================================

def bersihkan_angka(text):

    text = text.replace(",", ".")

    text = text.replace("O", "0")
    text = text.replace("o", "0")

    text = text.replace("I", "1")
    text = text.replace("l", "1")

    text = text.replace("S", "5")

    text = re.sub(r'[^0-9\.-]', '', text)

    return text

# ==========================================================
# FUNGSI MEMBACA ROI
# ==========================================================

def baca_roi(frame, roi):

    x, y, w, h = roi

    margin = 4

    x = max(0, x - margin)
    y = max(0, y - margin)

    w = w + margin * 2
    h = h + margin * 2

    crop = frame[y:y+h, x:x+w]

    crop = cv2.resize(
        crop,
        None,
        fx=4,
        fy=4,
        interpolation=cv2.INTER_CUBIC
    )

    gray = cv2.cvtColor(
        crop,
        cv2.COLOR_BGR2GRAY
    )

    gray = cv2.GaussianBlur(
        gray,
        (3,3),
        0
    )

    _, thresh = cv2.threshold(
        gray,
        0,
        255,
        cv2.THRESH_BINARY + cv2.THRESH_OTSU
    )

    hasil = reader.readtext(
        thresh,
        detail=0,
        allowlist="0123456789.-"
    )

    if len(hasil) == 0:
        return ""

    return bersihkan_angka(hasil[0])

def normalisasi_value(name, value):

    try:
        value = float(value)
    except:
        return value

    # ==========================
    # MW
    # ==========================

    if "MW" in name:

        # contoh:
        # 44 -> 4.4
        # 89 -> 8.9

        while value > MAX_MW:

            value /= 10

    # ==========================
    # MVAR
    # ==========================

    if "MVAR" in name:

        # contoh:
        # 3 -> 0.3
        # 15 -> 1.5
        # 30 -> 0.3

        while value > MAX_MVAR:

            value /= 10

    return round(value,2)

print("\n=== Monitoring dimulai ===\n")

try:

    wb = load_workbook(EXCEL_FILE)

except Exception:

    print("File Excel rusak atau tidak valid. Membuat file baru...")

    wb = Workbook()
    ws = wb.active
    ws.title = "LOG"

    ws.append([
        "Timestamp",
        "UNIT1_MW",
        "UNIT1_MVAR",
        "UNIT2_MW",
        "UNIT2_MVAR",
        "UNIT3_MW",
        "UNIT3_MVAR",
        "UNIT4_MW",
        "UNIT4_MVAR",
        "UNIT5_MW",
        "UNIT5_MVAR"
    ])

    wb.save(EXCEL_FILE)

ws = wb["LOG"]

last_values = {key: "" for key in ROIS.keys()}

last_reconnect = time.time()

last_excel = time.time()
SAVE_EXCEL = 5      # simpan Excel setiap 30 detik

while True:

    if time.time() - last_reconnect >= RECONNECT_INTERVAL:

        print("Reconnect stream...")

        cap.release()

        STREAM_URL = None

        while STREAM_URL is None:

            STREAM_URL = cari_live()

            if STREAM_URL is None:

                print("Menunggu live...")

                time.sleep(CHECK_LIVE_INTERVAL)

        cap = cv2.VideoCapture(STREAM_URL)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

        if cap.isOpened():
            last_reconnect = time.time()
        else:
            print("Tidak bisa membuka stream.")
            cap.release()
            time.sleep(5)

        continue

    with frame_lock:
        frame = None if latest_frame is None else latest_frame.copy()

    if frame is None:
        time.sleep(0.05)
        continue

    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")

    hasil = {}

    for name, roi in ROIS.items():

        value = baca_roi(frame, roi)

        if value != "":
            try:

                if "MW" in name or "MVAR" in name:
                    value = float(value) / 1000

                value = normalisasi_value(name, value)

            except:
                value = ""

        if value == "":
            value = last_values.get(name, "")

        hasil[name] = value
        last_values[name] = value


    # ======================================
    # Simpan JSON (cukup 1 kali)
    # ======================================

    with open(JSON_FILE, "w") as f:
        json.dump(
            {
                "timestamp": timestamp,
                "data": hasil
            },
            f,
            indent=4
        )

    print("=" * 50)
    print(timestamp)

    for k, v in hasil.items():
        print(f"{k:12} : {v}")

    ws.append([
        timestamp,

        hasil.get("UNIT1_MW",""),
        hasil.get("UNIT1_MVAR",""),

        hasil.get("UNIT2_MW",""),
        hasil.get("UNIT2_MVAR",""),

        hasil.get("UNIT3_MW",""),
        hasil.get("UNIT3_MVAR",""),

        hasil.get("UNIT4_MW",""),
        hasil.get("UNIT4_MVAR",""),

        hasil.get("UNIT5_MW",""),
        hasil.get("UNIT5_MVAR",""),
    ])

    if time.time() - last_excel >= SAVE_EXCEL:

        wb.save(EXCEL_FILE)

        print("💾 Excel disimpan")

        last_excel = time.time()

    time.sleep(INTERVAL)